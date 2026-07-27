"""Data engine: fetch → auto-map columns → normalise → derive metrics.

The workbook's exact columns are not assumed. `map_columns` fuzzy-matches
whatever headers arrive against a set of canonical roles, so re-ordering
columns, renaming them slightly, or adding new phases keeps working with
no code change. Everything downstream (KPIs, charts, health, risk) is then
computed from the canonical record, in the frontend, so filters stay live.
"""
from __future__ import annotations

import csv
import io
import os
import re
from datetime import date, datetime
from urllib.parse import urlencode

import requests
from dateutil import parser as dateparser

from config import (
    HEALTH_THRESHOLDS,
    PROJECT_ASSOCIATION,
    PROJECT_CLIENT,
    PROJECT_NAME,
    SHEET_ID,
    SHEET_TAB,
    STATUS_CLASSES,
    STATUS_LABELS,
    STATUS_PROGRESS,
)

# ── Canonical roles and the header keywords that map to them ──────────
# Order matters: earlier, more-specific roles win a header before a
# broader role can claim it (e.g. "Actual Completion" → actual, not end).
COLUMN_SYNONYMS: list[tuple[str, list[str]]] = [
    ("sr_no", ["sr no", "sr.", "s.no", "serial", "sno", "#", "id", "sl no"]),
    ("actual", ["actual completion", "actual", "completed on", "done on"]),
    ("start", ["start date", "start", "begin", "kickoff", "from date"]),
    ("end", ["end date", "due date", "due", "target date", "end", "deadline", "finish", "to date"]),
    ("meeting", ["meeting date", "meeting", "review date", "call date"]),
    ("tat", ["tat", "days", "duration", "turnaround", "effort", "timeline"]),
    ("qty", ["qty", "quantity", "count", "no.", "number of", "nos"]),
    ("progress", ["progress", "% complete", "percent", "completion %", "% done", "complete %"]),
    ("priority", ["priority", "urgency", "criticality"]),
    ("status", ["status", "state", "stage", "phase status"]),
    ("owner_oam", ["owner (oam)", "owner oam", "oam owner", "alphamed", "oam", "onealphamed"]),
    ("owner_hetero", ["owner (hetero)", "owner hetero", "hetero owner", "client owner", "hetero"]),
    ("reviewer", ["reviewer", "reviewed by", "approver", "qa by"]),
    ("owner", ["owner", "assignee", "responsible", "assigned to", "lead"]),
    ("dependencies", ["dependency", "dependencies", "depends on", "blocked by"]),
    ("approval", ["approval status", "approval", "approved", "sign off", "sign-off"]),
    ("category", ["category", "type", "workstream", "track"]),
    ("module", ["module", "component", "section"]),
    ("client", ["client", "customer", "account"]),
    ("association", ["association", "partner", "society", "body"]),
    ("version", ["version", "ver", "rev"]),
    ("comments", ["comments", "remarks", "notes", "comment", "remark"]),
    ("deliverable", ["deliverable", "task", "activity", "description", "work item", "milestone", "title", "name"]),
]

TODAY_OVERRIDE = os.environ.get("TRACKER_TODAY", "")  # for testing only


def _today() -> date:
    if TODAY_OVERRIDE:
        try:
            return dateparser.parse(TODAY_OVERRIDE, dayfirst=True).date()
        except Exception:
            pass
    return date.today()


# ── Fetch ─────────────────────────────────────────────────────────────

def fetch_sheet_rows() -> list[list[str]]:
    """Read the public sheet via the anonymous gviz CSV endpoint."""
    if not SHEET_ID:
        return []
    params = {"tqx": "out:csv"}
    if SHEET_TAB:
        params["sheet"] = SHEET_TAB
    url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/gviz/tq?{urlencode(params)}"
    try:
        r = requests.get(url, timeout=20)
        if r.status_code != 200:
            return []
        text = r.content.decode("utf-8", errors="replace")
        return list(csv.reader(io.StringIO(text)))
    except Exception:
        return []


def rows_from_excel(path: str) -> list[list[str]]:
    """Read the first sheet of an uploaded workbook into list-of-lists."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            return list(csv.reader(f))
    # xlsx / xlsm via openpyxl
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if c is None else str(c).strip() for c in row])
    wb.close()
    return rows


# ── Header handling ───────────────────────────────────────────────────

def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def find_header_row(rows: list[list[str]]) -> int:
    """Pick the row most likely to be the header (the one whose cells best
    match known column keywords). Handles a title/banner row above it."""
    best_idx, best_score = 0, -1
    all_kw = [kw for _, kws in COLUMN_SYNONYMS for kw in kws]
    for i, row in enumerate(rows[:8]):
        score = 0
        for cell in row:
            c = _norm(cell)
            if not c:
                continue
            if any(kw in c or c in kw for kw in all_kw):
                score += 1
        if score > best_score:
            best_idx, best_score = i, score
    return best_idx


def map_columns(headers: list[str]) -> dict[str, int]:
    """Map canonical role → column index by fuzzy keyword match.

    Each header is claimed by at most one role; each role by at most one
    header (best score wins). Falls back to the widest text column for the
    deliverable name if nothing matched it.
    """
    norm_headers = [_norm(h) for h in headers]
    used_cols: set[int] = set()
    mapping: dict[str, int] = {}

    for role, keywords in COLUMN_SYNONYMS:
        best_col, best_score = -1, 0
        for idx, h in enumerate(norm_headers):
            if idx in used_cols or not h:
                continue
            score = 0
            for kw in keywords:
                if h == kw:
                    score = max(score, 100)
                elif len(kw) >= 3 and (h.startswith(kw) or kw.startswith(h)):
                    score = max(score, 60)
                # word-boundary contains only — avoids "ver" matching
                # "deliVERable" or "id" matching "provided", etc.
                elif re.search(r"\b" + re.escape(kw) + r"\b", h):
                    score = max(score, 40)
            if score > best_score:
                best_col, best_score = idx, score
        if best_col >= 0:
            mapping[role] = best_col
            used_cols.add(best_col)

    # Guarantee a deliverable/name column: first unused non-empty header,
    # otherwise column 1 (right of an Sr-No style first column).
    if "deliverable" not in mapping:
        for idx, h in enumerate(norm_headers):
            if idx not in used_cols and h:
                mapping["deliverable"] = idx
                used_cols.add(idx)
                break
        else:
            mapping["deliverable"] = 1 if len(headers) > 1 else 0
    return mapping


def extract_phase(headers: list[str]) -> str:
    """Pull a "Phase X: ..." label out of any header cell (this workbook
    embeds it in the Sr-No header)."""
    for h in headers:
        m = re.search(r"(phase\s*[\dIVX]+.*)", h or "", re.IGNORECASE)
        if m:
            return m.group(1).strip()
    return ""


# ── Value parsing ─────────────────────────────────────────────────────

def parse_date(val: str) -> date | None:
    v = (val or "").strip()
    if not v or v.lower() in {"tbd", "tba", "na", "n/a", "-", "nil"}:
        return None
    try:
        return dateparser.parse(v, dayfirst=True, fuzzy=True).date()
    except Exception:
        return None


def status_class(raw: str) -> str:
    v = _norm(raw)
    for cls, values in STATUS_CLASSES.items():
        if v in values:
            return cls
    # partial contains match
    for cls, values in STATUS_CLASSES.items():
        for token in values:
            if token and token in v:
                return cls
    return "not_started"


def _to_int(val: str) -> int | None:
    m = re.search(r"-?\d+", (val or "").replace(",", ""))
    return int(m.group()) if m else None


# ── Normalisation ─────────────────────────────────────────────────────

def build_records(rows: list[list[str]]) -> dict:
    """Turn raw list-of-lists into the analytics payload the frontend uses."""
    today = _today()
    if not rows:
        return _empty_payload(today)

    hdr_idx = find_header_row(rows)
    headers = [h.strip() for h in rows[hdr_idx]]
    colmap = map_columns(headers)
    phase = extract_phase(headers) or "Phase 1"

    def cell(row: list[str], role: str) -> str:
        i = colmap.get(role)
        if i is None or i >= len(row):
            return ""
        return (row[i] or "").strip()

    records = []
    for ridx, row in enumerate(rows[hdr_idx + 1:]):
        name = cell(row, "deliverable")
        if not name and not any((c or "").strip() for c in row):
            continue
        if not name:
            continue

        start_d = parse_date(cell(row, "start"))
        end_d = parse_date(cell(row, "end"))
        actual_d = parse_date(cell(row, "actual"))
        meeting_d = parse_date(cell(row, "meeting"))
        raw_status = cell(row, "status")
        scls = status_class(raw_status)

        # Progress: explicit column wins, else implied by status.
        prog_raw = _to_int(cell(row, "progress"))
        progress = prog_raw if prog_raw is not None else STATUS_PROGRESS[scls]
        progress = max(0, min(100, progress))
        if scls == "done":
            progress = 100

        is_done = scls == "done"
        # Delay: overdue vs end date, or late actual completion.
        delay = 0
        if end_d:
            if is_done and actual_d:
                delay = max(0, (actual_d - end_d).days)
            elif not is_done:
                delay = max(0, (today - end_d).days)
        overdue = bool(end_d and not is_done and today > end_d)

        days_remaining = (end_d - today).days if (end_d and not is_done) else None
        upcoming = bool(days_remaining is not None and 0 <= days_remaining <= 7)

        health = "done" if is_done else _health_band(delay)

        owners = []
        for r in ("owner", "owner_oam", "owner_hetero", "reviewer"):
            v = cell(row, r)
            if v and v.lower() not in {"na", "n/a", "-"}:
                owners.append(v)

        records.append({
            "id": cell(row, "sr_no") or str(ridx + 1),
            "phase": phase,
            "deliverable": name,
            "qty": _to_int(cell(row, "qty")),
            "start": start_d.isoformat() if start_d else None,
            "end": end_d.isoformat() if end_d else None,
            "actual": actual_d.isoformat() if actual_d else None,
            "meeting": meeting_d.isoformat() if meeting_d else None,
            "start_label": start_d.strftime("%d %b %Y") if start_d else "",
            "end_label": end_d.strftime("%d %b %Y") if end_d else "",
            "tat": _to_int(cell(row, "tat")),
            "status": raw_status or STATUS_LABELS[scls],
            "status_class": scls,
            "status_label": STATUS_LABELS[scls],
            "priority": cell(row, "priority"),
            "owner_oam": cell(row, "owner_oam"),
            "owner_hetero": cell(row, "owner_hetero"),
            "reviewer": cell(row, "reviewer"),
            "owners": owners,
            "primary_owner": owners[0] if owners else "Unassigned",
            "dependencies": cell(row, "dependencies"),
            "approval": cell(row, "approval"),
            "category": cell(row, "category"),
            "module": cell(row, "module"),
            "client": cell(row, "client") or PROJECT_CLIENT,
            "association": cell(row, "association") or PROJECT_ASSOCIATION,
            "version": cell(row, "version"),
            "comments": cell(row, "comments"),
            "progress": progress,
            "days_remaining": days_remaining,
            "delay": delay,
            "overdue": overdue,
            "upcoming": upcoming,
            "health": health,
            "missing_owner": len(owners) == 0,
        })

    return {
        "ok": True,
        "project_name": PROJECT_NAME,
        "client": PROJECT_CLIENT,
        "association": PROJECT_ASSOCIATION,
        "phase": phase,
        "columns_detected": {k: headers[v] for k, v in colmap.items() if v < len(headers)},
        "records": records,
        "today": today.isoformat(),
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }


def _health_band(delay: int) -> str:
    for threshold, band in HEALTH_THRESHOLDS:
        if delay > threshold:
            return band
    return "green"


def _empty_payload(today: date) -> dict:
    return {
        "ok": False,
        "project_name": PROJECT_NAME,
        "client": PROJECT_CLIENT,
        "association": PROJECT_ASSOCIATION,
        "phase": "",
        "columns_detected": {},
        "records": [],
        "today": today.isoformat(),
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }
